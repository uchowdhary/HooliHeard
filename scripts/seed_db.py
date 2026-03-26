#!/usr/bin/env python3
"""Seed the database with insights from xlsx, JSON, or generate sample data."""

import json
import math
import re
import sys
import uuid
from datetime import date, datetime, timezone
from pathlib import Path

# Allow running from repo root or scripts/
ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT / "backend"))

from app.config import settings  # noqa: E402
from app.db import engine, SessionLocal, create_all  # noqa: E402
from app.models.insight import Insight  # noqa: E402

# --- File discovery ---
# Primary: two enriched xlsx files (Last30Days + 2MonthsPrior)
XLSX_LAST30 = ROOT / "data" / "output" / "Customer_Insights_Last30Days.xlsx"
XLSX_2MONTHS = ROOT / "data" / "output" / "Customer_Insights_2MonthsPrior.xlsx"
# Voice of Field weekly insights
VOF_XLSX = ROOT / "data" / "output" / "VoF_Weekly_Insights.xlsx"
# Fallback paths for Docker / k8s
if not XLSX_LAST30.exists():
    XLSX_LAST30 = Path("Customer_Insights_Last30Days.xlsx")
if not XLSX_2MONTHS.exists():
    XLSX_2MONTHS = Path("Customer_Insights_2MonthsPrior.xlsx")
if not VOF_XLSX.exists():
    VOF_XLSX = Path("VoF_Weekly_Insights.xlsx")

# Legacy fallbacks
XLSX_FILE = ROOT / "data" / "output" / "Insights_Combined_2026.xlsx"
INSIGHTS_JSON = ROOT / "data" / "output" / "insights.json"
if not INSIGHTS_JSON.exists():
    INSIGHTS_JSON = Path("insights.json")

# --- Priority scoring ---

STAGE_WEIGHTS = {
    "Closed Won": 10, "Negotiation": 8, "Negotiations": 8,
    "Technical Evaluation": 7, "Proposal": 7, "Capacity Review": 6,
    "Active Discussion / BMaaS": 5, "Active Customer": 5,
    "Discovery / Prospect": 3, "Discovery": 3, "Prospecting": 2, "New": 1,
    "Closed Lost": 1,
}

CATEGORY_URGENCY = {
    "Customer Requirements (Blocker)": 10,
    "CX Requirement": 8,
    "Capacity Issues": 8,
    "Issues": 7,
    "Loss Signal (Capacity)": 7,
    "Loss Signal (Commercial)": 7,
    "Loss Signal (Product Model Mismatch)": 7,
    "Loss Signal (No Response / Stale)": 5,
    "Loss Signal (Unknown)": 5,
    "Process / Operational Friction": 6,
    "Customer Requirements (Enhancement)": 5,
    "Capacity": 5,
    "Competition / Alternatives": 5,
    "Pricing / Terms": 4,
    "Education Gaps": 3,
    "GTM / Partnership": 3,
    "Product Fit / Scope": 6,
    "Success Pattern / Win Signal": 2,
    "Null": 1,
}

TIER_MULTIPLIERS = {
    "AI Enterprise": 3.0, "AI Lab": 2.5, "AI Platform": 2.0, "AI Native": 1.5,
    "Top Account": 3.0, "Strategic": 3.0, "Growth": 2.0, "Standard": 1.0,
}


# --- VoF category normalization ---
# Maps messy VoF categories to the canonical 13 categories
VOF_CATEGORY_MAP = {
    # Exact matches (case-insensitive handled in code)
    "capacity": "Capacity",
    "capacity ask": "Capacity",
    "capacity issues": "Capacity Issues",
    "capacity issues & customer requirements (blocker)": "Customer Requirements (Blocker)",
    "capacity issues - customer looking for capacity we can't support right now": "Capacity Issues",
    "capacity issues; customer requirements (blocker)": "Customer Requirements (Blocker)",
    "customer requirements": "Customer Requirements (Enhancement)",
    "customer requirements (blocker)": "Customer Requirements (Blocker)",
    "customer requirements (enhancemen) & pricing": "Customer Requirements (Enhancement)",
    "customer requirements (enhancement)": "Customer Requirements (Enhancement)",
    "customer requirements (enhancement); education gaps": "Customer Requirements (Enhancement)",
    "customer requirements (enhancement; capacity planning)": "Customer Requirements (Enhancement)",
    "customer requirements (potential blocker)": "Customer Requirements (Blocker)",
    "customer requirements – blocker": "Customer Requirements (Blocker)",
    "customer requirements – blocker (compliance)": "Customer Requirements (Blocker)",
    "customer requirements – blocker (scale/availability & timeline)": "Customer Requirements (Blocker)",
    "customer requirements – enhancement": "Customer Requirements (Enhancement)",
    "customer requirements – enhancement (commercial structure)": "Customer Requirements (Enhancement)",
    "customer requirements – enhancement (events/dc); education gaps (parquet guidance)": "Customer Requirements (Enhancement)",
    "customer requirements — blocker": "Customer Requirements (Blocker)",
    "customer requirements — enhancement": "Customer Requirements (Enhancement)",
    "education gap": "Education Gaps",
    "education gaps": "Education Gaps",
    "education gaps - customers don't know something exists": "Education Gaps",
    "education gaps; customer requirements (enhancement)": "Education Gaps",
    "issues": "Issues",
    "issues - these are technical issues customers faced": "Issues",
    "issues; customer requirements (enhancement)": "Issues",
    "issues; customer requirements (enhancement); capacity issues": "Issues",
    "issues; education gaps; customer requirements (enhancement)": "Issues",
    "null": "Null",
    "other": "Null",
    "partnership": "GTM / Partnership",
    "performance": "Issues",
    "pricing": "Pricing / Terms",
    "pricing & capacity": "Pricing / Terms",
    "pricing / terms": "Pricing / Terms",
    "pricing/terms": "Pricing / Terms",
}

# Maps VoF product areas to canonical 4
VOF_AREA_MAP = {
    "infra": "Infra",
    "cks": "CKS",
    "platform": "Platform",
    "ai services": "AI Services",
    "ai services (sunk/inference)": "AI Services",
    "w&b": "AI Services",
    "weights & biases": "AI Services",
    "multiple": "Infra",
    "other (cost)": "Infra",
    "other (term)": "Infra",
}

# Default subcategory by product area
VOF_SUBCATEGORY_DEFAULTS = {
    "Infra": "Compute",
    "CKS": "CKS",
    "Platform": "Console / API / Terraform",
    "AI Services": "Inference",
}

# VoF column name normalization — maps all header variants to canonical keys
VOF_COL_MAP = {
    "account": "account_name",
    "account name": "account_name",
    "account name (sfdc)": "account_name",
    "insight": "insight_text",
    "insights": "insight_text",
    "category": "insight_category",
    "categories": "insight_category",
    "insight category": "insight_category",
    "product area": "product_area",
    "product areas": "product_area",
    "product subcategory": "product_subcategory",
    "subcategory": "product_subcategory",
    "sales stage": "opportunity_stage",
    "sales stage (now)": "opportunity_stage",
    "opportunity $ amount": "opportunity_amount",
    "opportunity amount ($)": "opportunity_amount",
    "opportunity amount (usd)": "opportunity_amount",
    "opportunity $ amount": "opportunity_amount",
    "customer size": "use_case",
    "customer size (gpus / scale)": "gpu_types",
    "customer size / capacity signal": "gpu_types",
    "customer use case": "use_case",
    "gong call link": "source_link",
    "gong / sf link": "source_link",
    "gong links": "source_link",
    "sfdc / gong call link": "source_link",
    "gong link": "source_link",
    "gong call link": "source_link",
    "ent or native?": "icp",
}


# Normalize messy VoF stages to canonical 7 + Other
# Order: check substrings from most specific to least
_STAGE_NORMALIZE = [
    ("closed won", "Closed Won"),
    ("closed lost", "Closed Lost"),
    ("legal redline", "Legal Redlines"),
    ("negotiation", "Negotiations"),
    ("capacity review", "Capacity Review"),
    ("technical evaluation", "Technical Evaluation"),
    ("qualification", "Discovery"),
    ("discovery", "Discovery"),
    ("prospect", "Discovery"),
    ("pipeline", "Discovery"),
]


def _normalize_stage(raw: str) -> str:
    """Map a messy VoF opportunity stage to one of 7 canonical values or 'Other'."""
    low = raw.lower()
    if low in ("n/a", "no open opp", "no active opportunity", "", "none"):
        return "Other"
    for substring, canonical in _STAGE_NORMALIZE:
        if substring in low:
            return canonical
    if "current customer" in low or "active customer" in low:
        return "Closed Won"
    if "no open" in low or "no active" in low or "not found" in low or "former" in low:
        return "Other"
    return "Other"


def _parse_vof_sheet_date(sheet_name: str) -> date:
    """Parse a VoF sheet name like '3926' → 2026-03-09, '111725' → 2025-11-17."""
    name = sheet_name.strip()
    # Try to parse as MMDDYY or MDDYY
    if len(name) == 6:
        # MMDDYY
        m, d, y = int(name[:2]), int(name[2:4]), 2000 + int(name[4:6])
    elif len(name) == 5:
        # MDDYY
        m, d, y = int(name[:1]), int(name[1:3]), 2000 + int(name[3:5])
    elif len(name) == 4:
        # MDYY — e.g., 3926 → M=3, D=9, YY=26
        m, d, y = int(name[0]), int(name[1]), 2000 + int(name[2:4])
    else:
        return date.today()
    try:
        return date(y, m, d)
    except ValueError:
        return date.today()


def load_vof_xlsx(path: Path) -> list[dict]:
    """Load Voice of Field weekly insights from a multi-sheet xlsx file."""
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True)
    all_data = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        # Detect title row (first cell is long description, second cell is None)
        if rows[0][1] is None and len(str(rows[0][0] or "")) > 30:
            raw_headers = rows[1]
            data_rows = rows[2:]
        else:
            raw_headers = rows[0]
            data_rows = rows[1:]

        # Normalize headers to lowercase for mapping
        headers = [str(h).strip().lower() if h else "" for h in raw_headers]

        sheet_date = _parse_vof_sheet_date(sheet_name)

        for row in data_rows:
            # Skip empty rows
            if not any(cell for cell in row[:3]):
                continue

            record = {}
            for i, header in enumerate(headers):
                if not header:
                    continue
                key = VOF_COL_MAP.get(header, header)
                val = row[i] if i < len(row) else None
                if val is not None:
                    record[key] = val

            # Skip rows with no insight text
            if not record.get("insight_text"):
                continue

            # Normalize category
            raw_cat = str(record.get("insight_category") or "Null").strip().rstrip(".")
            record["insight_category"] = VOF_CATEGORY_MAP.get(raw_cat.lower(), "Null")

            # Normalize product area
            raw_area = str(record.get("product_area") or "Infra").strip()
            record["product_area"] = VOF_AREA_MAP.get(raw_area.lower(), "Infra")

            # Default subcategory if missing
            if not record.get("product_subcategory"):
                record["product_subcategory"] = VOF_SUBCATEGORY_DEFAULTS.get(
                    record["product_area"], "General"
                )

            # Normalize opportunity stage to canonical 7 + Other
            stage = record.get("opportunity_stage")
            if stage:
                record["opportunity_stage"] = _normalize_stage(str(stage).strip().rstrip("."))

            # Normalize opportunity amount
            opp = record.get("opportunity_amount")
            if opp is not None:
                try:
                    record["opportunity_amount"] = float(opp)
                except (ValueError, TypeError):
                    record["opportunity_amount"] = None

            # Normalize ICP from "ENT or Native?" column
            icp_raw = record.get("icp")
            if icp_raw:
                icp_str = str(icp_raw).strip().upper()
                if "ENT" in icp_str:
                    record["icp"] = "AI Enterprise"
                elif "NATIVE" in icp_str:
                    record["icp"] = "AI Native"
                else:
                    record["icp"] = None

            # Set VoF-specific fields
            record["date_of_record"] = sheet_date
            record["source_tool"] = "Gong"
            record["input_data_source"] = "Voice of Field"

            # Build dedup group key
            record["dedup_group_key"] = "|".join([
                str(record.get("account_name", "")),
                str(record.get("date_of_record", "")),
                str(record.get("product_area", "")),
                str(record.get("product_subcategory", "")),
                str(record.get("insight_category", "")),
            ])

            all_data.append(record)

    wb.close()
    return all_data


def compute_priority_score(row):
    """RICE-inspired priority score using log-scale for opportunity amount."""
    opp = row.get("opportunity_amount") or 0
    opp_score = min(10, max(1, math.log10(max(opp, 1)) - 2)) if opp > 0 else 1

    stage = row.get("opportunity_stage") or ""
    stage_score = STAGE_WEIGHTS.get(stage, 3)

    revenue = row.get("total_revenue") or 0
    won_count = row.get("closed_won_opp_count") or 0
    if revenue == 0 and won_count == 0:
        # v3 data has no revenue/won-count columns; use baseline so priority isn't artificially low
        engagement_score = 3
    else:
        engagement_score = min(10, 1 + (math.log10(max(revenue, 1)) - 2 if revenue > 100 else 0) + won_count * 0.5)

    cat = row.get("insight_category") or ""
    cat_score = CATEGORY_URGENCY.get(cat, 3)

    tier = row.get("icp") or row.get("account_priority_group") or "Standard"
    multiplier = TIER_MULTIPLIERS.get(tier, 1.0)

    raw = (opp_score * stage_score * engagement_score * cat_score * multiplier) / 100
    return round(raw, 2)


def compute_urgency_level(row):
    cat = row.get("insight_category") or ""
    score = CATEGORY_URGENCY.get(cat, 3)
    if score >= 8:
        return "High"
    elif score >= 5:
        return "Medium"
    return "Low"


def load_xlsx(path: Path) -> list[dict]:
    """Load insights from an xlsx file.

    Handles two formats:
      - Title row on row 1, headers on row 2 (new enriched files)
      - Headers on row 1 (legacy Insights_Combined_2026.xlsx)
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(path), read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))

    # Detect if row 0 is a title row (first cell is long/descriptive, second cell is None)
    if rows[0][1] is None and len(str(rows[0][0] or "")) > 30:
        # Title row present — headers are on row 1 (index 1)
        headers = rows[1]
        data_rows = rows[2:]
    else:
        headers = rows[0]
        data_rows = rows[1:]

    data = []

    # Column name mapping (xlsx header → internal key)
    COL_MAP = {
        "Account Name": "account_name",
        "Account Name (SFDC)": "account_name",
        "ICP": "icp",
        "Vertical": "vertical",
        "Stage": "opportunity_stage",
        "Opp Amount ($)": "opportunity_amount",
        "Use Case": "use_case",
        "GPU Types": "gpu_types",
        "Insight": "insight_text",
        "Product Area": "product_area",
        "Product Subcategory": "product_subcategory",
        "Insight Category": "insight_category",
        "Input Data Source": "input_data_source",
        "Source Tool": "source_tool",
        "Source Link": "source_link",
        "Date of Record": "date_of_record",
        "Period": "period",
        "Conversation Type": "conversation_type",
        "Role Present on Call": "role_present",
        "Comments": "comments",
        "Unique Insight Status": "unique_insight_status",
        "Account Priority Group": "account_priority_group",
        "Workloads": "workloads",
        "GPU Types / Quantities": "gpu_types",
        "Opportunity Stage": "opportunity_stage",
        "Opportunity Amount ($)": "opportunity_amount",
        "Competitors Mentioned": "competitors_mentioned",
        "Total Revenue": "total_revenue",
        "Most Recent Month w/ Revenue": "most_recent_revenue_month",
        "Closed Won Opp Count": "closed_won_opp_count",
    }

    for row in data_rows:
        record = {}
        for i, header in enumerate(headers):
            key = COL_MAP.get(header, header)
            val = row[i] if i < len(row) else None
            record[key] = val

        # Skip rows with no insight text
        if not record.get("insight_text"):
            continue

        # Normalize date
        d = record.get("date_of_record")
        if isinstance(d, datetime):
            record["date_of_record"] = d.date()
        elif isinstance(d, str):
            record["date_of_record"] = date.fromisoformat(d[:10])
        elif d is None:
            record["date_of_record"] = date.today()

        # Normalize numeric fields to float
        for num_field in ["opportunity_amount", "total_revenue"]:
            val = record.get(num_field)
            if val is not None:
                try:
                    record[num_field] = float(val)
                except (ValueError, TypeError):
                    record[num_field] = None

        # Normalize closed_won_opp_count to int
        cwoc = record.get("closed_won_opp_count")
        if cwoc is not None:
            try:
                record["closed_won_opp_count"] = int(cwoc)
            except (ValueError, TypeError):
                record["closed_won_opp_count"] = None

        # Build dedup group key
        record["dedup_group_key"] = "|".join([
            str(record.get("account_name", "")),
            str(record.get("date_of_record", "")),
            str(record.get("product_area", "")),
            str(record.get("product_subcategory", "")),
            str(record.get("insight_category", "")),
        ])

        data.append(record)

    wb.close()
    return data


def load_insights(data: list[dict], session) -> int:
    count = 0
    for row in data:
        priority = compute_priority_score(row)
        urgency = compute_urgency_level(row)

        d = row.get("date_of_record")
        if isinstance(d, str):
            d = date.fromisoformat(d[:10])
        elif isinstance(d, datetime):
            d = d.date()

        insight = Insight(
            id=uuid.uuid4(),
            account_name=row.get("account_name") or "Unknown",
            insight_text=row["insight_text"],
            product_area=row.get("product_area", "Unknown"),
            product_subcategory=row.get("product_subcategory", "General"),
            insight_category=row.get("insight_category", "Null"),
            input_data_source=row.get("input_data_source"),
            source_tool=row.get("source_tool", "unknown"),
            source_link=row.get("source_link"),
            role_present=row.get("role_present"),
            conversation_type=row.get("conversation_type"),
            date_of_record=d,
            comments=row.get("comments"),
            dedup_group_key=row.get("dedup_group_key"),
            unique_insight_status=row.get("unique_insight_status", "Key Record"),
            # Account enrichment
            icp=row.get("icp"),
            account_priority_group=row.get("account_priority_group"),
            vertical=row.get("vertical"),
            use_case=row.get("use_case"),
            workloads=row.get("workloads"),
            opportunity_stage=row.get("opportunity_stage"),
            opportunity_amount=row.get("opportunity_amount"),
            gpu_types=row.get("gpu_types"),
            competitors_mentioned=row.get("competitors_mentioned"),
            total_revenue=row.get("total_revenue"),
            most_recent_revenue_month=row.get("most_recent_revenue_month"),
            closed_won_opp_count=row.get("closed_won_opp_count"),
            # Computed
            priority_score=priority,
            urgency_level=urgency,
        )
        session.add(insight)
        count += 1
    session.commit()
    return count


def main():
    print("Creating tables...")
    create_all()

    session = SessionLocal()
    try:
        existing = session.query(Insight).count()
        if existing > 0:
            print(f"Database has {existing} insights. Clearing for re-seed...")
            session.query(Insight).delete()
            session.commit()

        # Prefer enriched xlsx files, then JSON fallback
        if XLSX_LAST30.exists() or XLSX_2MONTHS.exists():
            data = []
            for xlsx_path in [XLSX_LAST30, XLSX_2MONTHS]:
                if xlsx_path.exists():
                    print(f"Loading insights from {xlsx_path}")
                    data.extend(load_xlsx(xlsx_path))
        elif INSIGHTS_JSON.exists():
            print(f"Loading insights from {INSIGHTS_JSON}")
            with open(INSIGHTS_JSON) as f:
                data = json.load(f)
            if isinstance(data, dict) and "insights" in data:
                data = data["insights"]
        elif XLSX_FILE.exists():
            print(f"Loading insights from {XLSX_FILE}")
            data = load_xlsx(XLSX_FILE)
        else:
            print("No data file found. Cannot seed.")
            return

        count = load_insights(data, session)
        print(f"Seeded {count} insights from primary sources.")

        # Load Voice of Field data (additive)
        if VOF_XLSX.exists():
            print(f"Loading Voice of Field insights from {VOF_XLSX}")
            vof_data = load_vof_xlsx(VOF_XLSX)
            vof_count = load_insights(vof_data, session)
            count += vof_count
            print(f"Seeded {vof_count} VoF insights.")

        print(f"Total: {count} insights in the database.")
    finally:
        session.close()


if __name__ == "__main__":
    main()
